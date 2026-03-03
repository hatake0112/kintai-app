import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, date
import re

st.set_page_config(page_title="勤怠データ統合ツール", page_icon="📊", layout="wide")

st.title("📊 勤怠データ統合ツール")
st.markdown("各店舗の勤怠CSVファイルをアップロードすると、1つのExcelファイルに統合し、労務チェックを行います。")

with st.sidebar:
    st.header("⚙️ 設定")
    output_mode = st.radio(
        "Excel出力モード",
        ["すべて1シートに統合", "店舗ごとにシート分け"],
        index=0
    )
    add_store_column = st.checkbox("「店舗名」列を追加する", value=True)
    st.divider()
    st.header("📐 列名の設定")
    st.markdown("CSVの列名に合わせて入力してください")
    col_total_work = st.text_input("総労働時間の列名", value="総労働")
    col_break_time = st.text_input("休憩時間の列名", value="休憩時間")
    col_name = st.text_input("名前の列名", value="名前")


def normalize_name(name):
    if pd.isna(name):
        return ""
    s = str(name).strip()
    s = re.sub(r'[\s\u3000]+', '', s)
    return s


def read_csv_auto(file):
    encodings = ["utf-8", "cp932", "shift_jis", "utf-8-sig"]
    for enc in encodings:
        try:
            file.seek(0)
            df = pd.read_csv(file, encoding=enc)
            return df
        except (UnicodeDecodeError, Exception):
            file.seek(0)
            continue
    raise ValueError("文字コードを自動判定できませんでした")


def time_to_minutes(t):
    if pd.isna(t):
        return 0
    t = str(t).strip()
    if ":" in t:
        parts = t.split(":")
        try:
            h = int(parts[0])
            m = int(parts[1])
            return h * 60 + m
        except ValueError:
            return 0
    else:
        try:
            return int(float(t) * 60)
        except ValueError:
            return 0


def is_minor(birthday, check_date=None):
    if check_date is None:
        check_date = date.today()
    if pd.isna(birthday):
        return None
    try:
        if isinstance(birthday, str):
            birthday = birthday.strip()
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"]:
                try:
                    birthday = datetime.strptime(birthday, fmt).date()
                    break
                except ValueError:
                    continue
            else:
                return None
        elif isinstance(birthday, datetime):
            birthday = birthday.date()
        elif isinstance(birthday, pd.Timestamp):
            birthday = birthday.date()
        elif not isinstance(birthday, date):
            return None
        age = check_date.year - birthday.year - ((check_date.month, check_date.day) < (birthday.month, birthday.day))
        return age < 18
    except Exception:
        return None


def load_employee_data(excel_file):
    employees = {}
    xls = pd.ExcelFile(excel_file)
    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            name_col = None
            birthday_col = None
            header_row = None
            for i in range(min(5, len(df))):
                for j in range(len(df.columns)):
                    val = str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ""
                    if val == "名前" and name_col is None:
                        name_col = j
                        header_row = i
                    if "生年月日" in val and birthday_col is None:
                        birthday_col = j
            if name_col is None or birthday_col is None:
                continue
            for i in range(header_row + 1, len(df)):
                name_val = df.iloc[i, name_col]
                bday_val = df.iloc[i, birthday_col]
                if pd.isna(name_val) or str(name_val).strip() == "":
                    continue
                name_clean = normalize_name(name_val)
                name_clean = re.sub(r'\(.*?\)', '', name_clean)
                name_clean = re.sub(r'（.*?）', '', name_clean)
                if name_clean == "":
                    continue
                minor = is_minor(bday_val)
                if minor is not None:
                    employees[name_clean] = {"birthday": bday_val, "is_minor": minor, "original_name": str(name_val).strip()}
        except Exception:
            continue
    return employees


def check_labor(df, col_work, col_break, col_nm, employee_data):
    warnings = []
    warn_types = []
    minor_flags = []
    for _, row in df.iterrows():
        msgs = []
        types = []
        work_min = time_to_minutes(row.get(col_work, 0))
        break_min = time_to_minutes(row.get(col_break, 0))
        name_val = normalize_name(row.get(col_nm, ""))
        emp = employee_data.get(name_val, None)
        minor = emp["is_minor"] if emp else None
        if work_min > 360 and break_min < 30:
            msgs.append("⚠️ 休憩時間が30分以上確保されていません")
            types.append("break_short")
        if work_min > 480 and break_min < 60:
            msgs.append("🔴 休憩時間が足りません（8時間超は1時間以上必要）")
            types.append("break_lack")
        if work_min > 480:
            if minor is True:
                msgs.append("🚨 年少者が8時間を超えています！法令違反の可能性")
                types.append("minor_over8h")
            elif minor is False:
                pass
            else:
                msgs.append("🟡 8時間を超えています（従業員マスタに該当者なし・年少者確認必要）")
                types.append("over8h_unknown")
        warnings.append(" ／ ".join(msgs) if msgs else "")
        warn_types.append(",".join(types) if types else "")
        if minor is True:
            minor_flags.append("年少者")
        elif minor is False:
            minor_flags.append("")
        else:
            minor_flags.append("不明")
    df["⚠️ 労務チェック"] = warnings
    df["_warn_types"] = warn_types
    df["年少者判定"] = minor_flags
    return df


def create_report_excel(problem_df, col_work, col_break, col_nm):
    output = BytesIO()
    report_cols = []
    for c in ["店舗名", col_nm, "日付", col_work, col_break, "年少者判定", "⚠️ 労務チェック"]:
        if c in problem_df.columns:
            report_cols.append(c)
    report_df = problem_df[report_cols].copy()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        report_df.to_excel(writer, sheet_name="労務チェック結果", index=False)
        ws = writer.sheets["労務チェック結果"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        dark_red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        orange_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        check_col_idx = None
        for i, c in enumerate(report_cols):
            if c == "⚠️ 労務チェック":
                check_col_idx = i + 1
                break
        for row_idx in range(2, ws.max_row + 1):
            check_val = ws.cell(row=row_idx, column=check_col_idx).value if check_col_idx else ""
            check_str = str(check_val) if check_val else ""
            if "🚨" in check_str:
                fill = dark_red_fill
            elif "🔴" in check_str:
                fill = red_fill
            elif "🟡" in check_str:
                fill = yellow_fill
            elif "⚠️" in check_str:
                fill = orange_fill
            else:
                fill = None
            for col_idx in range(1, len(report_cols) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill
        for i, col in enumerate(report_cols, 1):
            max_len = len(str(col))
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=i).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 4, 50)
        today_str = datetime.now().strftime("%Y年%m月%d日")
        ws.sheet_properties.tabColor = "FF0000"
    output.seek(0)
    return output


if "dismissed" not in st.session_state:
    st.session_state.dismissed = set()

st.subheader("📁 ファイルアップロード")

col_up1, col_up2 = st.columns(2)
with col_up1:
    st.markdown("**① 従業員マスタ（Excelファイル）**")
    employee_file = st.file_uploader("クルー入社情報登録シート等", type=["xlsx", "xls"], key="emp")
with col_up2:
    st.markdown("**② 勤怠CSV（複数選択可）**")
    uploaded_files = st.file_uploader("各店舗の勤怠CSVファイル", type=["csv"], accept_multiple_files=True, key="csv")

employee_data = {}
if employee_file:
    employee_data = load_employee_data(employee_file)
    st.success(f"✅ 従業員マスタ読込完了: {len(employee_data)} 名")
    with st.expander("📋 読み込んだ従業員一覧", expanded=False):
        emp_list = []
        for k, v in employee_data.items():
            emp_list.append({"名前": v["original_name"], "生年月日": str(v["birthday"])[:10], "年少者": "✅ はい" if v["is_minor"] else "いいえ"})
        if emp_list:
            st.dataframe(pd.DataFrame(emp_list), use_container_width=True)
else:
    st.info("👆 従業員マスタをアップロードすると、年少者を自動判定します（なくても使えます）")

if uploaded_files:
    st.success(f"✅ {len(uploaded_files)} 件の勤怠CSVがアップロードされました")
    all_dataframes = {}
    errors = []
    for file in uploaded_files:
        try:
            df = read_csv_auto(file)
            store_name = file.name.replace(".csv", "").replace(".CSV", "")
            if add_store_column:
                df.insert(0, "店舗名", store_name)
            all_dataframes[store_name] = df
            st.write(f"**📁 {file.name}** — {len(df)} 行, {len(df.columns)} 列")
        except Exception as e:
            errors.append(f"❌ {file.name}: {e}")
    if errors:
        for err in errors:
            st.error(err)
    if all_dataframes:
        sample_df = list(all_dataframes.values())[0]
        has_work_col = col_total_work in sample_df.columns
        has_break_col = col_break_time in sample_df.columns
        has_name_col = col_name in sample_df.columns
        if not has_work_col or not has_break_col:
            st.warning(f"⚠️ CSVに該当する列名が見つかりません。左サイドバーの列名の設定を確認してください。CSVの列名一覧: {', '.join(sample_df.columns.tolist())}")
        else:
            for store_name, df in all_dataframes.items():
                all_dataframes[store_name] = check_labor(df, col_total_work, col_break_time, col_name, employee_data)
            st.divider()
            st.subheader("🚨 労務チェック結果")
            all_checked = pd.concat(all_dataframes.values(), ignore_index=True)
            problem_rows = all_checked[all_checked["⚠️ 労務チェック"] != ""].copy()
            problem_rows["_row_id"] = [f"{r.get('店舗名','')}__{r.get(col_name,'')}__{r.get('日付','')}" for _, r in problem_rows.iterrows()]
            problem_rows = problem_rows[~problem_rows["_row_id"].isin(st.session_state.dismissed)]
            if len(problem_rows) == 0:
                st.success("✅ すべてのデータに問題はありません！")
            else:
                st.error(f"⚠️ {len(problem_rows)} 件の問題が見つかりました")
                for idx, row in problem_rows.iterrows():
                    row_id = row["_row_id"]
                    warn_types = str(row.get("_warn_types", ""))
                    check_msg = row["⚠️ 労務チェック"]
                    store = row.get("店舗名", "")
                    name = row.get(col_name, "")
                    date_val = row.get("日付", "")
                    total = row.get(col_total_work, "")
                    brk = row.get(col_break_time, "")
                    minor_flag = row.get("年少者判定", "")
                    if "minor_over8h" in warn_types:
                        bg_color = "#ff9999"
                    elif "🔴" in str(check_msg):
                        bg_color = "#ffcccc"
                    elif "over8h_unknown" in warn_types:
                        bg_color = "#ffffcc"
                    elif "⚠️" in str(check_msg):
                        bg_color = "#fff3cd"
                    else:
                        bg_color = "#ffffff"
                    minor_badge = ""
                    if minor_flag == "年少者":
                        minor_badge = " <span style='background-color:#ff4444;color:white;padding:2px 8px;border-radius:4px;font-size:12px'>年少者</span>"
                    elif minor_flag == "不明":
                        minor_badge = " <span style='background-color:#999;color:white;padding:2px 8px;border-radius:4px;font-size:12px'>マスタ未登録</span>"
                    with st.container():
                        c1, c2 = st.columns([4, 1])
                        with c1:
                            st.markdown(f'<div style="background-color:{bg_color};padding:10px;border-radius:5px;margin-bottom:5px"><b>{store} / {name} / {date_val}</b>{minor_badge}<br>総労働: {total}　休憩: {brk}<br>{check_msg}</div>', unsafe_allow_html=True)
                        with c2:
                            if "over8h_unknown" in warn_types:
                                if "break_short" not in warn_types and "break_lack" not in warn_types:
                                    if st.button("✅ 年少者でない", key=f"ok_{row_id}"):
                                        st.session_state.dismissed.add(row_id)
                                        st.rerun()
                st.divider()
                st.subheader("📥 労務チェック結果のダウンロード")
                today_str = datetime.now().strftime("%Y%m%d")
                report_excel = create_report_excel(problem_rows, col_total_work, col_break_time, col_name)
                st.download_button(
                    label="📥 労務チェック結果をExcelでダウンロード",
                    data=report_excel,
                    file_name=f"労務チェック結果_{today_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                st.caption("※ 現在画面に表示されている問題（非表示にしたものを除く）がダウンロードされます")
        if st.session_state.dismissed:
            with st.expander(f"🗑️ 非表示にした項目（{len(st.session_state.dismissed)} 件）", expanded=False):
                for d in st.session_state.dismissed:
                    parts = d.split("__")
                    if len(parts) == 3:
                        st.write(f"- {parts[0]} / {parts[1]} / {parts[2]}")
                if st.button("🔄 すべて元に戻す"):
                    st.session_state.dismissed = set()
                    st.rerun()
        with st.expander("📋 全データプレビュー", expanded=False):
            for store_name, df in all_dataframes.items():
                st.subheader(store_name)
                show_df = df.drop(columns=["_warn_types"], errors="ignore")
                st.dataframe(show_df.head(20), use_container_width=True)
        def convert_to_excel():
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                if output_mode == "すべて1シートに統合":
                    combined = pd.concat(all_dataframes.values(), ignore_index=True)
                    combined = combined.drop(columns=["_warn_types"], errors="ignore")
                    combined.to_excel(writer, sheet_name="統合データ", index=False)
                else:
                    for store_name, df in all_dataframes.items():
                        sheet_name = store_name[:31]
                        show_df = df.drop(columns=["_warn_types"], errors="ignore")
                        show_df.to_excel(writer, sheet_name=sheet_name, index=False)
            output.seek(0)
            return output
        today = datetime.now().strftime("%Y%m%d")
        filename = f"勤怠データ統合_{today}.xlsx"
        st.divider()
        col1, col2 = st.columns([1, 3])
        with col1:
            excel_data = convert_to_excel()
            st.download_button(label="📥 統合データをExcelでダウンロード", data=excel_data, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with col2:
            if output_mode == "すべて1シートに統合":
                total_rows = sum(len(df) for df in all_dataframes.values())
                st.info(f"📊 統合結果: {len(all_dataframes)} 店舗 / {total_rows} 行")
            else:
                st.info(f"📊 {len(all_dataframes)} シートに分けて出力します")
else:
    st.info("👆 勤怠CSVファイルをドラッグ＆ドロップしてください")
st.divider()
st.caption("勤怠データ統合ツール v5.0 — 労務チェック結果ダウンロード機能付き")
